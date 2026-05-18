#!/usr/bin/env python3
"""
0427 확정 데이터 + 기준 엑셀 → 웹 대시보드 데이터 캐시 반영.

원칙:
- API 키는 사용하지 않는다. API 수집은 fetch_kosis.py / fetch_sgis.py가 담당한다.
- 0427_데이터에서 파싱 가능한 확정값은 source='local0427:*'로 기록한다.
- 기준 엑셀의 현장조사 시트는 별도 field-survey-meta.json으로 분리한다.
- 자동 파싱이 어려운 파일/지표는 data-gap-report.json에 명시한다.

실행:
  python import_0427_data.py
"""

import csv
import json
import math
import sys
from collections import defaultdict
from pathlib import Path
from datetime import datetime

sys.path.insert(0, str(Path(__file__).parent))
from lib_meta import load_full_meta, save_meta, SIGUN_CODES

if sys.stdout.encoding and sys.stdout.encoding.lower() != 'utf-8':
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        pass

try:
    from openpyxl import load_workbook
except ImportError:
    print('[ERROR] openpyxl 미설치. `pip install openpyxl`', file=sys.stderr)
    sys.exit(1)

try:
    import geopandas as gpd
    import pandas as pd
except ImportError:
    gpd = None
    pd = None


ROOT = Path(__file__).resolve().parents[1]
COURSE_ROOT = ROOT.parent
REF_XLSX = COURSE_ROOT / '공통지표_확정안+남양주 자율지표_0427.xlsx'
DATA_ROOT = COURSE_ROOT / '0427_데이터'
REGION_META_PATH = ROOT / 'dat' / 'region-meta.json'
INDICATOR_REF_PATH = ROOT / 'dat' / 'indicator-reference.json'
FIELD_SURVEY_PATH = ROOT / 'dat' / 'field-survey-meta.json'
GAP_REPORT_PATH = ROOT / 'dat' / 'data-gap-report.json'
SIGUN_GEOJSON_PATH = ROOT / 'dat' / 'gyeonggi-sigun.geojson'
DONG_GEOJSON_PATH = ROOT / 'dat' / 'gyeonggi-dong.geojson'


CITY_NAMES = {
    'namyangju': '남양주시',
    'gapyeong': '가평군',
    'yangpyeong': '양평군',
    'yeoju': '여주시',
    'icheon': '이천시',
    'anseong': '안성시',
    'pyeongtaek': '평택시',
    'hwaseong': '화성시',
    'osan': '오산시',
    'yongin': '용인시',
    'gwangju': '광주시',
    'hanam': '하남시',
    'yangju': '양주시',
    'pocheon': '포천시',
    'dongducheon': '동두천시',
}
CITY_NAME_TO_ID = {v: k for k, v in CITY_NAMES.items()}

# 엑셀 지표 ID와 현재 웹앱 내부 키의 호환 매핑.
EXCEL_TO_APP_KEY = {
    'R4': 'R5',  # 엑셀 R4 양호수질 하천 비율 → 앱 R5
    'R5': 'R6',  # 엑셀 R5 수변·생태쉼터 면적 → 앱 R6
    'W9': 'R4',  # 엑셀 W9 농촌체험 프로그램 → 앱 R4
}

APP_KEY_LABELS = {
    'L1': '인구증가율',
    'L2': '노령화지수',
    'L3': '인구순이동률',
    'L4': '생활SOC 충족지수',
    'W1': '고용률',
    'W2': '사업체수',
    'W3': '재정자립도',
    'W4': 'GRDP',
    'R1': '농촌환경 보전율',
    'R2': '토지이용 다양성 지수',
    'R3': '녹지율',
    'L5': '귀촌인 증감률',
    'L6': '3년 귀촌 규모 유지율',
    'W5': '농업 세대교체 수준',
    'W6': '청년 귀농 유입 비율',
    'W7': '친환경 인증 농가 비율',
    'R4': '인구 1천명당 농촌체험 프로그램',
    'R5': '양호수질 하천 비율',
    'R6': '수변·생태쉼터 면적',
    'R8': '국가유산',
    'W8': '서비스판매 종사자',
}

API_GAP_POLICIES = {
    'L3': {
        'status': 'partial',
        'reason': 'KOSIS DT_1B26001 표는 응답하지만 전체 차원 호출 시 40,000셀 제한에 걸림. 시군구 전입/전출 항목별 파라미터 확정 후 별도 파서 필요',
    },
    'L4': {
        'status': 'not_available',
        'reason': '기준 출처가 RAISE/농어촌서비스기준 실태조사로, 현재 KOSIS/SGIS Open API 수집 대상이 아님',
    },
    'W1': {
        'status': 'partial',
        'reason': 'KOSIS 경제활동인구조사 표는 있으나 현재 fetch_kosis.py 파라미터/공간단위 파서 미구현',
    },
    'W3': {
        'status': 'not_available',
        'reason': '기준 출처가 행안부 지방재정365로, KOSIS/SGIS Open API 대상이 아님',
    },
    'W4': {
        'status': 'partial',
        'reason': 'KOSIS 지역소득 표는 있으나 0427 로컬 파일이 1행 1열로 저장되어 파싱 불가. API 차원 파라미터 확정 필요',
    },
    'L5': {
        'status': 'partial',
        'reason': 'KOSIS 귀농귀촌 통계 표는 있으나 0427 로컬 파일이 1행 1열로 저장되어 파싱 불가. API 차원 파라미터 확정 필요',
    },
    'L6': {
        'status': 'partial',
        'reason': 'KOSIS 귀농귀촌 통계 표는 있으나 3년 전/현재 귀촌인수 항목 파라미터 확정 필요',
    },
    'R1': {
        'status': 'not_available',
        'reason': '기준 출처가 보호지역 DB/토지피복도 공간분석으로, KOSIS/SGIS 통계 API가 아닌 GIS 원천 처리 대상',
    },
    'R2': {
        'status': 'not_available',
        'reason': '기준 출처가 토지피복도 Shannon 다양성 산식으로, KOSIS/SGIS 통계 API가 아닌 GIS 원천 처리 대상',
    },
    'R3': {
        'status': 'not_available',
        'reason': '기준 출처가 토지피복도 녹지 면적 산식으로, KOSIS/SGIS 통계 API가 아닌 GIS 원천 처리 대상',
    },
    'R8': {
        'status': 'local_only',
        'reason': '국가유산은 0427 로컬/문화재청 계열 원천 확인 대상이며 KOSIS/SGIS 통계 API 수집 대상이 아님',
    },
}


gap_items = []
import_status = {}


def clean(value):
    if value is None:
        return ''
    return str(value).replace('\xa0', ' ').strip()


def cell(row, idx, default=''):
    return row[idx] if idx < len(row) else default


def to_number(value):
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if isinstance(value, float) and math.isnan(value):
            return None
        return value
    s = clean(value).replace(',', '')
    if not s or s in ('-', '—', '...', 'X'):
        return None
    try:
        return float(s) if '.' in s else int(s)
    except ValueError:
        return None


def rec(value, year, source, **extra):
    out = {'value': value, 'year': str(year), 'source': source}
    out.update(extra)
    return out


def add_gap(indicator, status, reason, source=None, scope='sigun'):
    gap_items.append({
        'indicator': indicator,
        'label': APP_KEY_LABELS.get(indicator, indicator),
        'status': status,
        'scope': scope,
        'reason': reason,
        'source': source or '',
    })


def app_key(excel_key):
    return EXCEL_TO_APP_KEY.get(excel_key, excel_key)


def ensure_manual(data, city_id):
    return data['sigun'][city_id].setdefault('manual', {})


def ensure_raw(data, city_id):
    return data['sigun'][city_id].setdefault('raw', {})


def ensure_computed(data, city_id):
    return data['sigun'][city_id].setdefault('computed', {})


def ensure_dong(data, adm_cd):
    data.setdefault('dong', {})
    data['dong'].setdefault(adm_cd, {'raw': {}, 'computed': {}, 'manual': {}})
    data['dong'][adm_cd].setdefault('raw', {})
    data['dong'][adm_cd].setdefault('computed', {})
    data['dong'][adm_cd].setdefault('manual', {})
    return data['dong'][adm_cd]


def read_xlsx(path, sheet=None):
    return load_workbook(path, read_only=True, data_only=True)[sheet] if sheet else load_workbook(path, read_only=True, data_only=True)


def build_dong_lookup():
    with open(DONG_GEOJSON_PATH, encoding='utf-8') as f:
        geo = json.load(f)
    by_name = {}
    for feature in geo.get('features', []):
        p = feature.get('properties', {})
        if p.get('city_id') != 'namyangju':
            continue
        name = clean(p.get('adm_nm'))
        if name:
            by_name[name] = p.get('adm_cd')
    return by_name


def read_reference_workbook():
    wb = load_workbook(REF_XLSX, read_only=True, data_only=True)
    indicators = {}

    # 총괄 시트
    ws = wb['1_총괄']
    for row in ws.iter_rows(values_only=True):
        if clean(cell(row, 0)) in ('ID', ''):
            continue
        key = clean(cell(row, 0))
        if key and key[0] in ('L', 'W', 'R'):
            mapped = app_key(key)
            indicators.setdefault(mapped, {
                'key': mapped,
                'source_key': key,
                'aliases': [],
            })
            if mapped != key:
                indicators[mapped]['aliases'].append(key)
            indicators[mapped].update({
                'category_label': clean(cell(row, 1)),
                'name': clean(cell(row, 2)),
                'formula': clean(cell(row, 3)),
                'spatial': clean(cell(row, 4)),
                'secured': clean(cell(row, 5)),
                'year': clean(cell(row, 6)),
            })

    # 상세 시트
    for sheet in ('2_삶터', '3_일터', '4_쉼터'):
        ws = wb[sheet]
        for row in ws.iter_rows(values_only=True):
            key = clean(cell(row, 1))
            if not key or key == 'ID' or key[0] not in ('L', 'W', 'R'):
                continue
            mapped = app_key(key)
            indicators.setdefault(mapped, {
                'key': mapped,
                'source_key': key,
                'aliases': [],
            })
            if mapped != key and key not in indicators[mapped]['aliases']:
                indicators[mapped]['aliases'].append(key)
            indicators[mapped].update({
                'type': clean(cell(row, 0)),
                'name': clean(cell(row, 2)) or indicators[mapped].get('name', ''),
                'formula': clean(cell(row, 3)) or indicators[mapped].get('formula', ''),
                'numerator': clean(cell(row, 4)),
                'denominator': clean(cell(row, 5)),
                'primary_source': clean(cell(row, 6)),
                'secondary_source': clean(cell(row, 7)),
                'url': clean(cell(row, 8)),
                'spatial': clean(cell(row, 9)) or indicators[mapped].get('spatial', ''),
            })

    field_survey = []
    ws = wb['7_현장조사']
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i == 1:
            continue
        values = [clean(v) for v in row[:8]]
        if not any(values):
            continue
        field_survey.append({
            'category_label': values[0],
            'type': values[1],
            'indicator_name': values[2],
            'purpose': values[3],
            'target': values[4],
            'sample_size': values[5],
            'method': values[6],
            'questions': values[7],
            'city_id': 'namyangju',
            'source': 'reference:7_현장조사',
        })

    # W6: 엑셀 총괄/상세에 부가설명이 비어 있어도 통계 표 열과 확정안을 맞춤
    w6 = indicators.get('W6')
    if isinstance(w6, dict) and not (w6.get('secondary_source') or '').strip().replace('-', ''):
        w6['secondary_source'] = '원천 표 연령 열 「30대이하」= 확정안 20~39세 구간'

    payload = {
        '_meta': {
            'source': str(REF_XLSX),
            'generated_at': datetime.now().isoformat(),
            'alias_policy': '웹앱 내부 키 유지, 엑셀 키는 aliases/source_key로 보존',
        },
        'indicators': indicators,
    }
    FIELD_SURVEY_PATH.write_text(json.dumps({
        '_meta': {
            'source': str(REF_XLSX),
            'sheet': '7_현장조사',
            'generated_at': datetime.now().isoformat(),
        },
        'items': field_survey,
    }, ensure_ascii=False, indent=2), encoding='utf-8')
    INDICATOR_REF_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding='utf-8')
    import_status['indicator_reference'] = {'status': 'ok', 'count': len(indicators)}
    import_status['field_survey'] = {'status': 'ok', 'count': len(field_survey)}
    return indicators, field_survey


def get_population(data, city_id):
    sigun = data.get('sigun', {}).get(city_id, {})
    for layer in ('raw', 'manual'):
        rec0 = sigun.get(layer, {}).get('population')
        if isinstance(rec0, dict) and rec0.get('value'):
            return to_number(rec0.get('value'))
    return None


def parse_housing(data):
    path = DATA_ROOT / '삶터' / '주택_현황_및_보급률_경기도.xlsx'
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    count = 0
    touched = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        city_id = CITY_NAME_TO_ID.get(clean(row[1]))
        item = clean(row[2])
        sub = clean(row[3])
        val = to_number(row[6] if len(row) > 6 else None)
        if not city_id or sub != '소계' or val is None:
            continue
        raw = ensure_raw(data, city_id)
        if item == '주택보급률 (%)':
            raw['housing_supply_rate'] = rec(val, 2023, 'local0427:housing')
            touched.add(city_id)
        elif item == '일반가구수 (가구)':
            raw['housing_households'] = rec(val, 2023, 'local0427:housing')
            touched.add(city_id)
        elif item == '주택수 (호)':
            raw['housing_count'] = rec(val, 2023, 'local0427:housing')
            touched.add(city_id)
    import_status['housing'] = {'status': 'ok', 'count': len(touched), 'path': str(path)}


def parse_medical(data):
    path = DATA_ROOT / '삶터' / '의료시설 수_경기도.xlsx'
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    counts = defaultdict(int)
    for row in ws.iter_rows(min_row=2, values_only=True):
        status = clean(row[18] if len(row) > 18 else '')
        if status != '영업중':
            continue
        address = clean(row[16] if len(row) > 16 else '')
        for cid, name in CITY_NAMES.items():
            if name in address:
                counts[cid] += 1
                break
    for cid, count in counts.items():
        ensure_raw(data, cid)['medical_facility_count'] = rec(count, 2026, 'local0427:medical_facilities')
    import_status['medical_facilities'] = {'status': 'ok', 'count': len(counts), 'path': str(path)}


def parse_experience_farms(data, dong_lookup):
    path = DATA_ROOT / '일터' / 'W9_남양주체험농장정보.xlsx'
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    by_dong = defaultdict(int)
    total = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        dong_name = clean(row[0])
        farm_name = clean(row[1])
        if not dong_name or not farm_name:
            continue
        by_dong[dong_name] += 1
        total += 1

    pop = get_population(data, 'namyangju')
    ensure_raw(data, 'namyangju')['rural_experience_farm_count'] = rec(total, 2024, 'local0427:experience_farms')
    if pop:
        ensure_computed(data, 'namyangju')['R4_experience_programs_per_1000'] = {
            'value': round(total / pop * 1000, 3),
            'unit': '건/천명',
            'formula': '농촌체험농장 수 / 주민등록인구 × 1,000',
            'inputs': {'farm_count': total, 'population': pop},
            'source': 'local0427:experience_farms',
        }

    for dong_name, count in by_dong.items():
        adm_cd = dong_lookup.get(dong_name)
        if not adm_cd:
            add_gap('R4', 'partial', f'체험농장 읍면명 매칭 실패: {dong_name}', str(path), 'dong')
            continue
        ensure_dong(data, adm_cd)['raw']['rural_experience_farm_count'] = rec(count, 2024, 'local0427:experience_farms')
    import_status['experience_farms'] = {'status': 'ok', 'count': total, 'path': str(path)}


def parse_parks(data, dong_lookup):
    path = DATA_ROOT / '쉼터' / 'R5_남양주시_도시공원정보.csv'
    rows = []
    for enc in ('utf-8-sig', 'cp949'):
        try:
            with open(path, encoding=enc, newline='') as f:
                rows = list(csv.DictReader(f))
            break
        except UnicodeDecodeError:
            continue
    total_area = 0
    total_count = 0
    by_dong = defaultdict(lambda: {'count': 0, 'area': 0})
    for row in rows:
        area = to_number(row.get('공원면적')) or 0
        total_area += area
        total_count += 1
        addr = clean(row.get('소재지도로명주소')) + ' ' + clean(row.get('소재지지번주소'))
        for dong_name in dong_lookup:
            if dong_name in addr:
                by_dong[dong_name]['count'] += 1
                by_dong[dong_name]['area'] += area
                break

    pop = get_population(data, 'namyangju')
    raw = ensure_raw(data, 'namyangju')
    raw['urban_park_count'] = rec(total_count, 2024, 'local0427:urban_parks')
    raw['urban_park_area'] = rec(round(total_area, 2), 2024, 'local0427:urban_parks')
    if pop:
        ensure_computed(data, 'namyangju')['R6_park_area_per_1000'] = {
            'value': round(total_area / pop * 1000, 2),
            'unit': '㎡/천명',
            'formula': '도시공원 면적 / 주민등록인구 × 1,000',
            'inputs': {'urban_park_area': round(total_area, 2), 'population': pop},
            'source': 'local0427:urban_parks',
        }

    for dong_name, vals in by_dong.items():
        adm_cd = dong_lookup.get(dong_name)
        if not adm_cd:
            continue
        dong = ensure_dong(data, adm_cd)
        dong['raw']['urban_park_count'] = rec(vals['count'], 2024, 'local0427:urban_parks')
        dong['raw']['urban_park_area'] = rec(round(vals['area'], 2), 2024, 'local0427:urban_parks')
    import_status['urban_parks'] = {'status': 'ok', 'count': total_count, 'path': str(path)}


def parse_agri_managers(data, dong_lookup):
    path = DATA_ROOT / '일터' / '농업경영체 현황_남양주.xlsx'
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    total = young = 0
    by_dong = defaultdict(lambda: {'total': 0, 'young': 0})
    for row in ws.iter_rows(min_row=3, values_only=True):
        dong_name = clean(row[0])
        age = clean(row[1])
        if dong_name in ('총계', '합계', '계'):
            continue
        manager_count = (to_number(row[2]) or 0) + (to_number(row[3]) or 0)
        if not dong_name or manager_count <= 0:
            continue
        total += manager_count
        by_dong[dong_name]['total'] += manager_count
        is_young = age.startswith('30세') or age.startswith('35세') or age.startswith('20세')
        if is_young:
            young += manager_count
            by_dong[dong_name]['young'] += manager_count

    raw = ensure_raw(data, 'namyangju')
    raw['agri_manager_total'] = rec(total, 2024, 'local0427:agri_managers')
    raw['agri_manager_young'] = rec(young, 2024, 'local0427:agri_managers')
    if total:
        ensure_computed(data, 'namyangju')['W5_agri_young_manager_ratio'] = {
            'value': round(young / total * 100, 2),
            'unit': '%',
            'formula': '39세 이하 농업경영주 / 전체 농업경영주 × 100',
            'inputs': {'young_managers': young, 'total_managers': total},
            'source': 'local0427:agri_managers',
        }

    for dong_name, vals in by_dong.items():
        adm_cd = dong_lookup.get(dong_name)
        if not adm_cd:
            add_gap('W5', 'partial', f'농업경영체 읍면명 매칭 실패: {dong_name}', str(path), 'dong')
            continue
        dong = ensure_dong(data, adm_cd)
        dong['raw']['agri_manager_total'] = rec(vals['total'], 2024, 'local0427:agri_managers')
        dong['raw']['agri_manager_young'] = rec(vals['young'], 2024, 'local0427:agri_managers')
        if vals['total']:
            dong['computed']['W5_agri_young_manager_ratio'] = {
                'value': round(vals['young'] / vals['total'] * 100, 2),
                'unit': '%',
                'formula': '39세 이하 농업경영주 / 전체 농업경영주 × 100',
                'inputs': vals,
                'source': 'local0427:agri_managers',
            }
    import_status['agri_managers'] = {'status': 'ok', 'count': total, 'path': str(path)}


def parse_return_farm(data):
    path = DATA_ROOT / '일터' / '귀농인현황_남양주.xlsx'
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    current_region = ''
    total = young = None
    for row in ws.iter_rows(min_row=3, values_only=True):
        if clean(row[0]):
            current_region = clean(row[0])
        item = clean(row[1])
        if current_region == '남양주시' and '귀농가구원수' in item:
            total = to_number(row[2])
            # 확정안 W6: 20~39세 귀농가구원 비중. 원천 표 열 명은 「30대이하」(연령대 구간 표기).
            young = to_number(row[3])
            break
    if total is None:
        add_gap('W6', 'local_parse_failed', '귀농인현황_남양주에서 남양주시 행을 찾지 못함', str(path))
        return
    raw = ensure_raw(data, 'namyangju')
    raw['return_farm_members'] = rec(total, 2024, 'local0427:return_farm')
    raw['return_farm_young_members'] = rec(young or 0, 2024, 'local0427:return_farm')
    if total:
        ensure_computed(data, 'namyangju')['W6_young_return_farm_ratio'] = {
            'value': round((young or 0) / total * 100, 2),
            'unit': '%',
            'formula': '20~39세 귀농가구원수(원천 표 「30대이하」) / 전체 귀농가구원수 × 100',
            'inputs': {'young_return_farm_members': young or 0, 'return_farm_members': total},
            'source': 'local0427:return_farm',
        }
    import_status['return_farm'] = {'status': 'ok', 'count': 1, 'path': str(path)}


def parse_eco_farm(data):
    path = DATA_ROOT / '일터' / '친환경인증농가현황_남양주.xlsx'
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    current_region = ''
    farms = None
    for row in ws.iter_rows(min_row=3, values_only=True):
        if clean(row[1]):
            current_region = clean(row[1])
        item = clean(row[2])
        if '남양주시' in current_region and '농가수' in item:
            farms = to_number(row[7])
            break
    if farms is None:
        add_gap('W7', 'local_parse_failed', '친환경인증농가현황에서 남양주시 농가수 행을 찾지 못함', str(path))
        return
    raw = ensure_raw(data, 'namyangju')
    raw['eco_certified_farms'] = rec(farms, 2024, 'local0427:eco_certified_farms')
    denominator = None
    farm_rec = raw.get('farm_cnt')
    if isinstance(farm_rec, dict):
        denominator = to_number(farm_rec.get('value'))
    if not denominator:
        manager_rec = raw.get('agri_manager_total')
        if isinstance(manager_rec, dict):
            denominator = to_number(manager_rec.get('value'))
    if denominator:
        ensure_computed(data, 'namyangju')['W7_eco_certified_farm_ratio'] = {
            'value': round(farms / denominator * 100, 2),
            'unit': '%',
            'formula': '친환경 인증 농가수 / 전체 농가수 × 100',
            'inputs': {'eco_certified_farms': farms, 'farm_denominator': denominator},
            'source': 'local0427:eco_certified_farms',
        }
    import_status['eco_certified_farms'] = {'status': 'ok', 'count': 1, 'path': str(path)}


def parse_jobs(data):
    path = DATA_ROOT / '일터' / '직업별_취업자_남양주시.xlsx'
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    latest = None
    for row in ws.iter_rows(min_row=4, values_only=True):
        period = clean(row[0])
        if period:
            latest = row
    if not latest:
        add_gap('W8', 'local_parse_failed', '직업별 취업자 남양주시 최신 반기 행 없음', str(path))
        return
    total = to_number(latest[1])
    service_sales = to_number(latest[7])
    if service_sales is not None:
        ensure_raw(data, 'namyangju')['service_sales_workers_local'] = rec(service_sales * 1000, 2023, 'local0427:jobs')
        ensure_computed(data, 'namyangju')['W8_service_sales_workers'] = {
            'value': int(service_sales * 1000),
            'unit': '명',
            'formula': '지역별고용조사 서비스판매종사자 취업자수(천명) × 1,000',
            'inputs': {'service_sales_workers_thousand': service_sales, 'total_workers_thousand': total},
            'source': 'local0427:jobs',
        }
    import_status['jobs_namyangju'] = {'status': 'ok', 'count': 1, 'path': str(path)}


def parse_water_quality(data):
    if gpd is None:
        add_gap('R5', 'not_available', 'geopandas 미설치로 수질 SHP 공간조인 불가', scope='sigun')
        return
    shp_dir = DATA_ROOT / '쉼터' / 'R4_2024_기후에너지환경부_물환경 수질측정망 정보(SHP)'
    shp_files = sorted(shp_dir.glob('*.shp'))
    if not shp_files:
        add_gap('R5', 'local_missing', '수질측정망 SHP 파일 없음', str(shp_dir))
        return
    sigun = gpd.read_file(SIGUN_GEOJSON_PATH).to_crs('EPSG:4326')
    frames = []
    for path in shp_files:
        gdf = gpd.read_file(path).to_crs('EPSG:4326')
        gdf = gdf[gdf['TYPE'].astype(str).str.contains('하천', na=False)].copy()
        frames.append(gdf[['ptNo', 'ptNm', 'itemBod', 'geometry']])
    points = pd.concat(frames, ignore_index=True)
    points = gpd.GeoDataFrame(points, geometry='geometry', crs='EPSG:4326')
    joined = gpd.sjoin(points, sigun[['id', 'geometry']], how='inner', predicate='within')
    if joined.empty:
        add_gap('R5', 'local_parse_failed', '수질측정망 지점이 15개 시군 폴리곤과 공간조인되지 않음', str(shp_dir))
        return
    count = 0
    for city_id, group in joined.groupby('id'):
        by_point = group.groupby('ptNo')['itemBod'].mean()
        if len(by_point) == 0:
            continue
        good = int((by_point <= 3.0).sum())  # BOD 3mg/L 이하를 Ⅱ등급 이상 대리지표로 사용
        total = int(len(by_point))
        rate = round(good / total * 100, 2)
        raw = ensure_raw(data, city_id)
        raw['good_water_points'] = rec(good, 2024, 'local0427:water_quality')
        raw['water_quality_points'] = rec(total, 2024, 'local0427:water_quality')
        ensure_computed(data, city_id)['R5_good_water_rate'] = {
            'value': rate,
            'unit': '%',
            'formula': 'BOD 평균 3mg/L 이하 측정지점 수 / 전체 하천 측정지점 수 × 100',
            'inputs': {'good_points': good, 'total_points': total},
            'source': 'local0427:water_quality',
        }
        count += 1
    import_status['water_quality'] = {'status': 'ok', 'count': count, 'path': str(shp_dir)}


def record_unparsed_known_files():
    known = [
        ('L5', DATA_ROOT / '삶터' / 'L5L6_시도별_시군별__·_성별_귀촌인_20260424004831_분석(전년_대비_증감,증감률).xlsx',
         '현재 파일이 데이터 시트 1행 1열만 포함해 자동 파싱 불가. KOSIS 원표/API 재조회 필요'),
        ('L6', DATA_ROOT / '삶터' / 'L5L6_시도별_시군별__·_성별_귀촌인_20260424004831_분석(전년_대비_증감,증감률).xlsx',
         '현재 파일이 데이터 시트 1행 1열만 포함해 자동 파싱 불가. KOSIS 원표/API 재조회 필요'),
        ('W4', DATA_ROOT / '일터' / 'W3_GRDP_시_군_구__20260424005846_.xlsx',
         '현재 파일이 데이터 시트 1행 1열만 포함해 자동 파싱 불가. KOSIS 원표/API 재조회 필요'),
        ('R8', DATA_ROOT / '쉼터' / 'R8_경기도_국가유산.xlsx',
         '현재 파일이 데이터 시트 1행 1열만 포함해 자동 파싱 불가. 경기도 통계표 재다운로드/API 확인 필요'),
    ]
    for indicator, path, reason in known:
        if path.exists():
            policy = API_GAP_POLICIES.get(indicator, {})
            add_gap(indicator, policy.get('status', 'local_parse_failed'), policy.get('reason', reason), str(path))


def build_gap_report(data, indicators, field_survey):
    field_names = {item['indicator_name'] for item in field_survey}
    for key in APP_KEY_LABELS:
        if any(g['indicator'] == key for g in gap_items):
            continue
        has_value = False
        for city in data.get('sigun', {}).values():
            for layer in ('raw', 'computed', 'manual'):
                if any(key in field or APP_KEY_LABELS[key] in field for field in city.get(layer, {})):
                    has_value = True
                    break
            if has_value:
                break
        if not has_value:
            if APP_KEY_LABELS[key] in field_names:
                add_gap(key, 'field_survey', '기준 엑셀 7_현장조사 시트의 담당자 입력/현장조사 대상')
            elif key in API_GAP_POLICIES:
                policy = API_GAP_POLICIES[key]
                add_gap(key, policy['status'], policy['reason'])
            else:
                add_gap(key, 'not_collected', 'region-meta에 확정값 없음. API 또는 0427 원천 재확인 필요')

    payload = {
        '_meta': {
            'generated_at': datetime.now().isoformat(),
            'source': 'reference workbook + local0427 import + current API coverage',
        },
        'import_status': import_status,
        'items': gap_items,
    }
    GAP_REPORT_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding='utf-8')


def update_coverage(data, field_survey):
    data.setdefault('_meta', {})
    data['_meta']['fetched_at'] = datetime.now().isoformat()
    data['_meta'].setdefault('tables', {})['local0427_import'] = {
        'status': 'ok',
        'generated_at': datetime.now().isoformat(),
        'items': import_status,
    }
    data['_meta']['coverage'] = {
        'sigun': sum(1 for d in data.get('sigun', {}).values() if d.get('raw') or d.get('manual') or d.get('computed')),
        'dong': len(data.get('dong', {})),
        'field_survey': len(field_survey),
    }


def main():
    if not REF_XLSX.exists():
        raise FileNotFoundError(REF_XLSX)
    if not DATA_ROOT.exists():
        raise FileNotFoundError(DATA_ROOT)

    print('[1/5] 기준 엑셀 지표/현장조사 메타 생성')
    indicators, field_survey = read_reference_workbook()
    data = load_full_meta()
    dong_lookup = build_dong_lookup()

    print('[2/5] 0427 로컬 확정 데이터 반영')
    parse_housing(data)
    parse_medical(data)
    parse_experience_farms(data, dong_lookup)
    parse_parks(data, dong_lookup)
    parse_agri_managers(data, dong_lookup)
    parse_return_farm(data)
    parse_eco_farm(data)
    parse_jobs(data)
    parse_water_quality(data)
    record_unparsed_known_files()

    print('[3/5] 갭 리포트 생성')
    build_gap_report(data, indicators, field_survey)

    print('[4/5] region-meta coverage 갱신')
    update_coverage(data, field_survey)
    save_meta(data)

    print('[5/5] 완료')
    print(f'  indicator-reference: {INDICATOR_REF_PATH}')
    print(f'  field-survey-meta:   {FIELD_SURVEY_PATH}')
    print(f'  gap-report:          {GAP_REPORT_PATH}')


if __name__ == '__main__':
    main()
