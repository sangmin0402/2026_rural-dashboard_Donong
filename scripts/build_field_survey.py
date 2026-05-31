#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
현장조사 가상데이터(xlsx) -> namyangju-field-survey.json 재현 가능 빌더.

배경
----
2026-05-31 현장조사 가상데이터(한유하)는 남양주 9개 농촌 읍면(6읍·3면)에 대한
현장조사 raw 4종(농가조사 / 귀촌정착 / 체험텃밭 / SOC체크)을 담고 있다.
이 스크립트는 raw 행을 읍면 단위로 group-by 하여 대시보드가 바로 읽을 수 있는
읍면별 지표값(CANON 키: W5 W7 L4 L6 W6 W9 R6 R7)으로 사전 집계한다.

산식(계획서 §2-1)
------------------
  W5 농업세대교체(%)   = count(청년경영주_여부=='Y') / n * 100            [농가조사]
  W7 친환경인증농가(%) = count(친환경_인증_여부=='Y') / n * 100          [농가조사]
  L4 생활SOC충족지수   = SOC체크 L4_SOC충족지수 (읍면 1행, 그대로)        [SOC체크]
  L6 3년귀촌정착률(%)  = (전입연도<=2023 중 2026주소유지=='Y') / 대상 * 100 [귀촌정착]
  W6 청년귀농유입(%)   = sum(20_39세_가구원수) / sum(가구원수) * 100      [귀촌정착]
  W9 농촌체험(건/천명) = sum(연간프로그램_횟수) / 읍면인구 * 1000         [체험텃밭]
  R6 도시텃밭수용(%)   = sum(수용가능인원[도시텃밭]) / 읍면인구 * 100      [체험텃밭]
  R7 주말농원활성화(%) = sum(운영구획[주말농원]) / sum(분양가능구획[주말농원]) * 100 [체험텃밭]

인구 분모는 region-meta.json(dong.raw.tot_ppltn)에서 읽는다. 없으면 mock 인구,
둘 다 없으면 value=null + _pending=true.

사용법
------
  python build_field_survey.py \
    --xlsx "../../0531작업/현장조사 가상데이터.xlsx" \
    --out  "../dat/simulation/namyangju-field-survey.json" \
    --pop-source "../dat/region-meta.json"

결정론적: 동일 입력 -> 동일 출력. Date.now() 등 비결정 요소 없음(generated_at는 인자/고정).
"""
import argparse
import json
import os
import sys

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("openpyxl 필요: pip install -r requirements.txt")

# 읍면명 -> (adm_cd, cluster) : namyangju-dong-mock.json 과 동일 코드/클러스터
EUP_META = {
    "와부읍":   ("31130110", "transit"),
    "화도읍":   ("31130111", "transit"),
    "진접읍":   ("31130120", "transit"),
    "진건읍":   ("31130140", "transit"),
    "오남읍":   ("31130150", "transit"),
    "퇴계원읍": ("31130160", "transit"),
    "별내면":   ("31130310", "rural"),
    "수동면":   ("31130340", "rural"),
    "조안면":   ("31130350", "rural"),
}
EUP_ORDER = list(EUP_META.keys())


def read_sheet(wb, name):
    """시트를 dict 행 리스트로 (헤더=1행)."""
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    out = []
    for r in rows[1:]:
        if all(c is None for c in r):
            continue
        out.append({header[i]: r[i] for i in range(len(header))})
    return out


def find_sheet(wb, *keywords):
    """이름에 keyword 가 포함된 첫 시트명."""
    for nm in wb.sheetnames:
        if any(k in nm for k in keywords):
            return nm
    return None


def num(v, default=0.0):
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def yn(v):
    return str(v).strip().upper() == "Y"


def r1(x):
    return None if x is None else round(x, 1)


def load_population(pop_source):
    """region-meta.json -> {adm_cd: tot_ppltn}."""
    pops = {}
    if pop_source and os.path.exists(pop_source):
        meta = json.load(open(pop_source, encoding="utf-8"))
        for code, d in meta.get("dong", {}).items():
            v = d.get("raw", {}).get("tot_ppltn", {}).get("value")
            if v:
                pops[code] = num(v, None)
    return pops


def group_by_eup(rows, key="읍면"):
    g = {e: [] for e in EUP_ORDER}
    for row in rows:
        e = str(row.get(key, "")).strip()
        if e in g:
            g[e].append(row)
    return g


def build(xlsx, pop_source):
    wb = load_workbook(xlsx, data_only=True)

    s_farm = find_sheet(wb, "농가")
    s_settle = find_sheet(wb, "귀촌")
    s_exp = find_sheet(wb, "체험", "텃밭")
    s_soc = find_sheet(wb, "SOC", "SCO")

    farm = group_by_eup(read_sheet(wb, s_farm)) if s_farm else {}
    settle = group_by_eup(read_sheet(wb, s_settle)) if s_settle else {}
    exp = group_by_eup(read_sheet(wb, s_exp)) if s_exp else {}
    soc_rows = read_sheet(wb, s_soc) if s_soc else []
    soc = {str(r.get("읍면", "")).strip(): r for r in soc_rows}

    pops = load_population(pop_source)
    eups = {}

    for name in EUP_ORDER:
        adm_cd, cluster = EUP_META[name]
        pop = pops.get(adm_cd)
        ind = {}

        # --- 농가조사: W5, W7 ---
        frows = farm.get(name, [])
        nf = len(frows)
        if nf:
            young = sum(1 for r in frows if yn(r.get("청년경영주_여부")))
            eco = sum(1 for r in frows if yn(r.get("친환경_인증_여부")))
            ind["W5"] = {"value": r1(young / nf * 100), "unit": "%", "label": "농업 세대교체 수준",
                         "source": "field-survey", "n": nf,
                         "formula": "청년경영주_여부=='Y' 비율"}
            ind["W7"] = {"value": r1(eco / nf * 100), "unit": "%", "label": "친환경 인증 농가 비율",
                         "source": "field-survey", "n": nf,
                         "formula": "친환경_인증_여부=='Y' 비율"}

        # --- SOC체크: L4 ---
        srow = soc.get(name)
        if srow is not None:
            l4 = srow.get("L4_SOC충족지수")
            if l4 is None and srow.get("총항목수"):
                l4 = num(srow.get("보유항목수")) / num(srow.get("총항목수")) * 100
            ind["L4"] = {"value": r1(num(l4)), "unit": "점", "label": "생활SOC 충족지수",
                         "source": "field-survey", "n": 1,
                         "formula": "보유항목수/총항목수*100 (12개 생활SOC)"}

        # --- 귀촌정착: L6, W6 ---
        srows = settle.get(name, [])
        cohort = [r for r in srows if num(r.get("전입연도"), 9999) <= 2023]
        if cohort:
            stayed = sum(1 for r in cohort if yn(r.get("2026년_주소유지_여부")))
            ind["L6"] = {"value": r1(stayed / len(cohort) * 100), "unit": "%",
                         "label": "귀촌인 3년 정착률", "source": "field-survey", "n": len(cohort),
                         "formula": "전입<=2023 중 2026 주소유지=='Y' 비율"}
        if srows:
            ymem = sum(num(r.get("20_39세_가구원수")) for r in srows)
            tmem = sum(num(r.get("가구원수")) for r in srows)
            if tmem:
                ind["W6"] = {"value": r1(ymem / tmem * 100), "unit": "%",
                             "label": "청년 귀농 유입 비율(20~39세)", "source": "field-survey",
                             "n": len(srows), "formula": "sum(20_39세)/sum(가구원수)*100"}

        # --- 체험텃밭: W9, R6, R7 ---
        erows = exp.get(name, [])
        if erows:
            prog = sum(num(r.get("연간프로그램_횟수")) for r in erows)
            if pop:
                ind["W9"] = {"value": r1(prog / pop * 1000), "unit": "건/천명",
                             "label": "인구 1천명당 농촌체험 프로그램 운영 건수",
                             "source": "field-survey", "n": len(erows),
                             "formula": "sum(연간프로그램_횟수)/인구*1000"}
            else:
                ind["W9"] = {"value": None, "unit": "건/천명",
                             "label": "인구 1천명당 농촌체험 프로그램 운영 건수",
                             "source": "field-survey", "n": len(erows), "_pending": True,
                             "formula": "인구 분모 미확보"}

            garden = [r for r in erows if "도시텃밭" in str(r.get("시설유형", ""))]
            cap = sum(num(r.get("수용가능인원")) for r in garden)
            if pop and garden:
                ind["R6"] = {"value": r1(cap / pop * 100), "unit": "%",
                             "label": "도시텃밭 체험 수용 가능 비율", "source": "field-survey",
                             "n": len(garden), "formula": "sum(수용가능인원[도시텃밭])/인구*100"}

            weekend = [r for r in erows if "주말농원" in str(r.get("시설유형", ""))]
            div = sum(num(r.get("주말농원_분양가능구획")) for r in weekend)
            run = sum(num(r.get("주말농원_운영구획")) for r in weekend)
            if div:
                ind["R7"] = {"value": r1(run / div * 100), "unit": "%",
                             "label": "주말농원 활성화율", "source": "field-survey",
                             "n": len(weekend),
                             "formula": "sum(운영구획)/sum(분양가능구획)*100 (현장조사 정의)"}

        eups[name] = {"adm_cd": adm_cd, "cluster": cluster,
                      "population": pop, "indicators": ind}

    return {
        "_meta": {
            "version": "2026-05-31",
            "source": "0531작업/현장조사 가상데이터.xlsx (가상 현장조사, 한유하)",
            "scope": "namyangju",
            "unit": "읍면",
            "eup_count": len(eups),
            "indicators": ["W5", "W7", "L4", "L6", "W6", "W9", "R6", "R7"],
            "id_scheme": "CANON (= xlsx 확정안). mock 파일과 R4/R5/R6 키 의미가 다름 — app.js 변환 레이어 경유.",
            "population_source": "region-meta.json dong.raw.tot_ppltn",
            "note": "가상(시뮬레이션) 현장조사 데이터. 실제 수집 전 UI·집계 검증용. R7은 현장조사 대체정의(분양 대비 운영구획)로 확정안 정의와 다를 수 있음.",
        },
        "eups": eups,
    }


def main():
    here = os.path.dirname(os.path.abspath(__file__))
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=os.path.join(here, "..", "..", "0531작업", "현장조사 가상데이터.xlsx"))
    ap.add_argument("--out", default=os.path.join(here, "..", "dat", "simulation", "namyangju-field-survey.json"))
    ap.add_argument("--pop-source", default=os.path.join(here, "..", "dat", "region-meta.json"))
    args = ap.parse_args()

    data = build(args.xlsx, args.pop_source)
    os.makedirs(os.path.dirname(args.out), exist_ok=True)
    with open(args.out, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    # 요약 출력
    print(f"[OK] {args.out}")
    for nm, e in data["eups"].items():
        ks = ",".join(e["indicators"].keys())
        print(f"  {nm} (pop={e['population']}): {ks}")


if __name__ == "__main__":
    main()
